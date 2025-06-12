from google.cloud import vision
from google.oauth2 import service_account
import os
from datetime import datetime
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import io
import concurrent.futures
from pathlib import Path
import logging
from typing import List, Dict

class OCRProcessor:
    def __init__(self, credentials_path):
        self._validate_paths(credentials_path)
        self.credentials = self._load_credentials(credentials_path)
        self.client = vision.ImageAnnotatorClient(credentials=self.credentials)

    def _validate_paths(self, credentials_path: str) -> None:
        cred_path = Path(credentials_path)
        if not cred_path.exists():
            raise FileNotFoundError(f"Credentials không tồn tại: {credentials_path}")
        if not cred_path.suffix == '.json':
            raise ValueError("Credentials phải là file JSON")

    def _load_credentials(self, credentials_path):
        try:
            if not os.path.exists(credentials_path):
                raise FileNotFoundError(f"Không tìm thấy file credentials: {credentials_path}")
            return service_account.Credentials.from_service_account_file(credentials_path)
        except Exception as e:
            raise Exception(f"Lỗi khi đọc credentials: {str(e)}")

    def process_single_image(self, image_path):
        """Process single image and return result"""
        try:
            # Phân tích tên file để lấy thông tin crop
            image_name = os.path.basename(image_path)
            crop_info = None
            if image_name.startswith('crop_'):
                try:
                    # Giả sử tên file có format: crop_XXX_className.png
                    crop_number = int(image_name.split('_')[1])
                    # Đọc file JSON chứa thông tin crop gốc
                    json_path = os.path.join(os.path.dirname(image_path), "processed_results.json")
                    if os.path.exists(json_path):
                        with open(json_path, 'r', encoding='utf-8') as f:
                            all_crops = json.load(f)
                            # Tìm thông tin crop tương ứng
                            crop_info = next((crop for crop in all_crops if crop['index'] == crop_number), None)
                except Exception as e:
                    print(f"Warning: Could not parse crop info from filename: {e}")

            with open(image_path, 'rb') as image_file:
                content = image_file.read()
            
            image = vision.Image(content=content)
            response = self.client.text_detection(
                image=image,
                image_context={"language_hints": ["vi", "vi-VN"]}
            )

            # Lấy kích thước ảnh
            img = Image.open(image_path)
            width, height = img.size

            # Chuẩn bị kết quả với thông tin crop
            result = {
                "image_name": image_name,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "image_size": {"width": width, "height": height},
                "original_crop_coords": crop_info['coords'] if crop_info else None,
                "full_text": "",
                "text_blocks": []
            }

            if response.error.message:
                print(f"Warning - API error: {response.error.message}")
                return result

            if response.text_annotations:
                result["full_text"] = response.text_annotations[0].description
                for text in response.text_annotations[1:]:
                    # Lấy 4 đỉnh của bounding box
                    vertices = [(vertex.x, vertex.y) for vertex in text.bounding_poly.vertices]
                    
                    # Tính điểm trung tâm của box
                    center_x = sum(v[0] for v in vertices) / 4
                    center_y = sum(v[1] for v in vertices) / 4
                    
                    # Thêm thông tin tọa độ gốc nếu có
                    if crop_info:
                        original_x1, original_y1, _, _ = crop_info['coords']
                        # Điều chỉnh tọa độ tương đối về tọa độ tuyệt đối
                        adjusted_vertices = [
                            (x + original_x1, y + original_y1) for x, y in vertices
                        ]
                        adjusted_center = {
                            "x": center_x + original_x1,
                            "y": center_y + original_y1
                        }
                    else:
                        adjusted_vertices = vertices
                        adjusted_center = {"x": center_x, "y": center_y}
                    
                    result["text_blocks"].append({
                        "text": text.description,
                        "position": {
                            "vertices": adjusted_vertices,
                            "center": adjusted_center,
                            "relative_vertices": vertices  # Giữ lại tọa độ tương đối
                        }
                    })

            return result

        except Exception as e:
            print(f"Error processing {os.path.basename(image_path)}: {str(e)}")
            return None

    def create_excel_report(self, results, output_excel_path, images_dir):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "OCR Results"

            headers = ["Image", "Image Name", "Text Content", "Coordinates"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            for idx, result in enumerate(results, 2):
                image_path = os.path.join(images_dir, result["image_name"])
                
                # Xử lý hình ảnh - với nhiều lần thử
                max_retries = 3
                for attempt in range(max_retries):
                    try:
                        if os.path.exists(image_path):
                            img = Image.open(image_path)
                            img = img.convert('RGB')
                            
                            # Resize với tỷ lệ cố định
                            target_size = (75, 75)
                            img = img.resize(target_size, Image.Resampling.LANCZOS)
                            
                            img_byte_arr = io.BytesIO()
                            img.save(img_byte_arr, format='PNG', optimize=True)
                            img_byte_arr.seek(0)
                            
                            xl_image = XLImage(img_byte_arr)
                            xl_image.width = 75
                            xl_image.height = 75
                            
                            ws.row_dimensions[idx].height = 60
                            cell = ws.cell(row=idx, column=1)
                            ws.add_image(xl_image, cell.coordinate)
                            break
                    except Exception as e:
                        if attempt == max_retries - 1:
                            print(f"Failed to add image after {max_retries} attempts: {result['image_name']}")
                        continue

                # Add text data
                ws.cell(row=idx, column=2, value=result["image_name"])
                ws.cell(row=idx, column=3, value=result.get("full_text", ""))

                # Format coordinates
                if result["text_blocks"]:
                    block = result["text_blocks"][0]
                    vertices = block["position"]["vertices"]
                    center = block["position"]["center"]
                    coordinates = (
                        f"Top-left: ({vertices[0][0]}, {vertices[0][1]})\n"
                        f"Top-right: ({vertices[1][0]}, {vertices[1][1]})\n"
                        f"Bottom-right: ({vertices[2][0]}, {vertices[2][1]})\n"
                        f"Bottom-left: ({vertices[3][0]}, {vertices[3][1]})\n"
                        f"Center: ({center['x']:.1f}, {center['y']:.1f})"
                    )
                else:
                    # Nếu không có text, lấy tọa độ từ kích thước ảnh
                    w = result["image_size"]["width"]
                    h = result["image_size"]["height"]
                    coordinates = (
                        f"Top-left: (0, 0)\n"
                        f"Top-right: ({w}, 0)\n"
                        f"Bottom-right: ({w}, {h})\n"
                        f"Bottom-left: (0, {h})\n"
                        f"Center: ({w/2:.1f}, {h/2:.1f})"
                    )
                ws.cell(row=idx, column=4, value=coordinates)

            # Adjust column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 35

            wb.save(output_excel_path)
            print(f"Excel report created successfully: {output_excel_path}")

        except Exception as e:
            raise Exception(f"Error creating Excel report: {str(e)}")

    def process_directory(self, input_dir: str, output_dir: str) -> int:
        try:
            # Validate directories
            input_path = Path(input_dir)
            output_path = Path(output_dir)
            
            if not input_path.exists():
                raise FileNotFoundError(f"Input directory không tồn tại: {input_dir}")
            
            output_path.mkdir(parents=True, exist_ok=True)
            
            # Filter image files
            image_files = [f for f in input_path.glob('crop_*.png')]
            if not image_files:
                raise ValueError(f"Không tìm thấy ảnh trong {input_dir}")

            # Process images in parallel
            with concurrent.futures.ThreadPoolExecutor(max_workers=4) as executor:
                future_to_image = {
                    executor.submit(self.process_single_image, str(img_path)): img_path
                    for img_path in image_files
                }
                
                all_results = []
                total = len(future_to_image)
                
                for idx, future in enumerate(concurrent.futures.as_completed(future_to_image), 1):
                    img_path = future_to_image[future]
                    try:
                        result = future.result()
                        if result:
                            all_results.append(result)
                        print(f"\rProgress: {(idx/total)*100:.1f}% ({idx}/{total})", end="")
                    except Exception as e:
                        logging.error(f"Error processing {img_path}: {str(e)}")

            # Save results
            if all_results:
                self._save_results(all_results, output_path, input_path)
            
            return len(all_results)

        except Exception as e:
            logging.error(f"Error in process_directory: {str(e)}")
            raise

    def _save_results(self, results: List[Dict], output_path: Path, input_path: Path) -> None:
        """Save results to JSON and Excel"""
        # Save JSON
        json_path = output_path / "json_results.json"
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

        # Create Excel report
        excel_path = output_path / "ocr_summary.xlsx"
        self.create_excel_report(results, str(excel_path), str(input_path))

def main():
    try:
        base_dir = r"E:\WORK\project\OCR\Recognition_OCR\data_results"
        credentials_path = os.path.join(base_dir, "credentials.json")
        input_dir = os.path.join(base_dir, "imgs")
        output_dir = os.path.join(base_dir, "ocr_output")

        print("=== BẮT ĐẦU QUÁ TRÌNH OCR ===")
        start_time = datetime.now()

        processor = OCRProcessor(credentials_path)
        processed_count = processor.process_directory(input_dir, output_dir)

        duration = (datetime.now() - start_time).total_seconds()
        print(f"\nĐã xử lý {processed_count} ảnh")
        print(f"Thời gian: {duration:.2f} giây")
        print(f"Kết quả được lưu tại: {output_dir}")

    except Exception as e:
        print(f"\nLỗi: {str(e)}")
    finally:
        print("\n=== KẾT THÚC ===")

if __name__ == "__main__":
    main()