import os
from openpyxl import load_workbook, Workbook
from pathlib import Path

def merge_excel_files(input_dir: str, output_file: str):
    print(f"Reading Excel files from: {input_dir}")
    
    # Create new workbook
    merged_wb = Workbook()
    merged_sheet = merged_wb.active
    merged_sheet.title = "Merged Results"
    
    # Set headers
    merged_sheet['A1'] = "STT"
    merged_sheet['B1'] = "Image"
    merged_sheet['C1'] = "Extracted Text"
    
    # Set column widths
    merged_sheet.column_dimensions['A'].width = 10
    merged_sheet.column_dimensions['B'].width = 30
    merged_sheet.column_dimensions['C'].width = 50
    
    current_row = 2  # Start from row 2 (after header)
    
    # Get and sort Excel files
    excel_files = [f for f in os.listdir(input_dir) if f.endswith('.xlsx')]
    excel_files.sort(key=lambda x: int(Path(x).stem) if Path(x).stem.isdigit() else float('inf'))
    
    print(f"Found {len(excel_files)} Excel files")
    
    for excel_file in excel_files:
        input_path = os.path.join(input_dir, excel_file)
        print(f"Processing: {excel_file}")
        
        try:
            wb = load_workbook(input_path)
            sheet = wb.active
            
            # Copy data from each row (skip header)
            for row_idx in range(2, sheet.max_row + 1):
                # Add sequence number
                merged_sheet[f'A{current_row}'] = current_row - 1
                
                # Copy images safely
                try:
                    if sheet._images:
                        img_index = row_idx - 2  # Adjust index for images
                        if img_index < len(sheet._images):
                            img = sheet._images[img_index]
                            # Calculate new anchor position
                            img.anchor._from.row = current_row - 1
                            img.anchor._from.col = 1  # Column B
                            merged_sheet.add_image(img)
                except Exception as img_err:
                    print(f"Warning: Could not copy image in row {row_idx}: {str(img_err)}")
                
                # Copy text
                text_value = sheet[f'B{row_idx}'].value
                merged_sheet[f'C{current_row}'] = text_value
                
                # Set row height
                merged_sheet.row_dimensions[current_row].height = 100  # Fixed height for consistency
                
                current_row += 1
                
        except Exception as e:
            print(f"Error processing {excel_file}: {str(e)}")
            continue
    
    print(f"Saving merged file to: {output_file}")
    merged_wb.save(output_file)
    print("Merge completed successfully!")

def main():
    base_dir = r"E:\WORK\project\OCR\Recognition_OCR\make_label"
    input_dir = os.path.join(base_dir, "output_excel_RNN")
    output_file = os.path.join(base_dir, "merged_results.xlsx")
    
    merge_excel_files(input_dir, output_file)

if __name__ == "__main__":
    main()