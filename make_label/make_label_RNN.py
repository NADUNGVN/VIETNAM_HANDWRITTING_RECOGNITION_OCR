import os
from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image
import cv2
from ultralytics import YOLO
from google.cloud import vision
from google.oauth2 import service_account

class ImageTextExtractor:
    def __init__(self, yolo_model_path: str, google_credentials_path: str):
        # Khởi tạo YOLO model
        self.yolo_model = YOLO(yolo_model_path)
        self.yolo_model.to('cpu')
        
        # Khởi tạo Google Vision client
        credentials = service_account.Credentials.from_service_account_file(google_credentials_path)
        self.vision_client = vision.ImageAnnotatorClient(credentials=credentials)
        self.temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temp")
        os.makedirs(self.temp_dir, exist_ok=True)

    def process_image(self, image_path: str) -> List[Dict]:
        try:
            # Đọc ảnh
            img = cv2.imread(image_path)
            if img is None:
                raise ValueError(f"Không thể đọc ảnh: {image_path}")

            # Phát hiện vùng text bằng YOLO
            results = self.yolo_model(img, conf=0.3, iou=0.45, device='cpu')
            if not results or len(results) == 0:
                return []

            detected_texts = []
            boxes = results[0].boxes

            # Xử lý từng vùng text
            for i, box in enumerate(boxes):
                x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                
                # Cắt vùng ảnh
                cropped_img = img[y1:y2, x1:x2]
                temp_path = os.path.join(self.temp_dir, f"temp_crop_{i}.png")
                cv2.imwrite(temp_path, cropped_img)

                # OCR bằng Google Vision
                with open(temp_path, 'rb') as image_file:
                    content = image_file.read()
                
                vision_image = vision.Image(content=content)
                response = self.vision_client.text_detection(
                    image=vision_image,
                    image_context={"language_hints": ["vi"]}
                )

                text = response.text_annotations[0].description if response.text_annotations else ""
                detected_texts.append({
                    'crop_image': temp_path,
                    'text': text
                })

            return detected_texts

        except Exception as e:
            print(f"Error in process_image: {str(e)}")
            return []

    def __del__(self):
        # Cleanup temp files when object is destroyed
        if hasattr(self, 'temp_dir') and os.path.exists(self.temp_dir):
            for file in os.listdir(self.temp_dir):
                try:
                    os.remove(os.path.join(self.temp_dir, file))
                except Exception as e:
                    print(f"Error removing temp file: {str(e)}")
            try:
                os.rmdir(self.temp_dir)
            except Exception as e:
                print(f"Error removing temp directory: {str(e)}")

    def create_excel_report(self, results: List[Dict], output_path: str):
        wb = Workbook()
        ws = wb.active
        ws.title = "OCR Results"

        # Set cột headers
        ws['A1'] = "Image"
        ws['B1'] = "Extracted Text"

        # Set độ rộng cột
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 50

        # Thêm dữ liệu
        for idx, result in enumerate(results, 2):
            # Thêm ảnh
            img = Image.open(result['crop_image'])
            # Resize ảnh để vừa với cell Excel
            max_height = 100
            width, height = img.size
            new_width = int((max_height/height) * width)
            img = img.resize((new_width, max_height))
            
            xl_image = XLImage(result['crop_image'])
            xl_image.width = new_width
            xl_image.height = max_height
            
            ws.row_dimensions[idx].height = max_height
            ws.add_image(xl_image, f'A{idx}')

            # Thêm text
            ws[f'B{idx}'] = result['text']

        wb.save(output_path)

def main():
    try:
        # Đường dẫn mới
        base_dir = r"E:\WORK\project\OCR\Recognition_OCR\make_label"
        model_path = os.path.join(base_dir, "..", "model", "best_yolo12n_v2_5_6_314img.pt")
        credentials_path = os.path.join(base_dir, "..", "data_results", "credentials.json")
        
        # Tạo thư mục output nếu chưa tồn tại
        output_dir = os.path.join(base_dir, "output_excel_RNN")
        os.makedirs(output_dir, exist_ok=True)
        
        input_image = os.path.join(base_dir, "input_RNN", "19.png")  
        output_excel = os.path.join(output_dir, "19.xlsx")

        print("Initializing extractor...")
        extractor = ImageTextExtractor(model_path, credentials_path)

        print(f"Processing image: {input_image}")
        results = extractor.process_image(input_image)

        print("Creating Excel report...")
        extractor.create_excel_report(results, output_excel)

        print(f"Done! Results saved to: {output_excel}")

    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    main()